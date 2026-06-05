import os
from io import BytesIO
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any
from fastapi import HTTPException
from weasyprint import HTML, CSS


class ExportService:
    @staticmethod
    def get_pdf_styles() -> str:
        """
        Retorna los estilos CSS compartidos para que el PDF se vea sumamente premium y oficial.
        """
        return """
            @page {
                size: letter;
                margin: 1.5cm;
                @bottom-right {
                    content: "Página " counter(page) " de " counter(pages);
                    font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
                    font-size: 8pt;
                    color: #718096;
                }
                @bottom-left {
                    content: "ERP Multisucursal - CONFIDENCIAL";
                    font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
                    font-size: 8pt;
                    color: #718096;
                    font-weight: bold;
                }
            }
            body {
                font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
                color: #2D3748;
                font-size: 9pt;
                line-height: 1.4;
            }
            .header-container {
                border-bottom: 2px solid #3182CE;
                padding-bottom: 12px;
                margin-bottom: 20px;
            }
            .header-table {
                width: 100%;
                border-collapse: collapse;
            }
            .company-name {
                font-size: 16pt;
                font-weight: bold;
                color: #1A365D;
            }
            .report-title {
                font-size: 13pt;
                font-weight: bold;
                color: #3182CE;
                text-transform: uppercase;
                margin-top: 4px;
            }
            .meta-info {
                text-align: right;
                font-size: 8pt;
                color: #4A5568;
            }
            .meta-info strong {
                color: #2D3748;
            }
            .table-report {
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
            }
            .table-report th {
                background-color: #2B6CB0;
                color: #FFFFFF;
                font-weight: bold;
                text-transform: uppercase;
                font-size: 8pt;
                padding: 6px 8px;
                border: 1px solid #2B6CB0;
                text-align: left;
            }
            .table-report td {
                padding: 6px 8px;
                border-bottom: 1px solid #E2E8F0;
            }
            .table-report tr:nth-child(even) td {
                background-color: #F7FAFC;
            }
            .text-right {
                text-align: right;
            }
            .text-center {
                text-align: center;
            }
            .font-bold {
                font-weight: bold;
            }
            .badge {
                display: inline-block;
                padding: 2px 6px;
                border-radius: 4px;
                font-size: 7.5pt;
                font-weight: bold;
            }
            .badge-ingreso { background-color: #E6FFFA; color: #00A389; }
            .badge-salida { background-color: #FFF5F5; color: #E53E3E; }
            .badge-traslado { background-color: #EBF8FF; color: #2B6CB0; }
            .badge-ajuste { background-color: #FEFCBF; color: #B7791F; }
            
            .summary-section {
                margin-top: 20px;
                background-color: #F7FAFC;
                border: 1px solid #E2E8F0;
                border-radius: 6px;
                padding: 12px;
                width: 40%;
                margin-left: auto;
            }
            .summary-table {
                width: 100%;
                border-collapse: collapse;
            }
            .summary-table td {
                padding: 4px 0;
            }
        """

    @staticmethod
    def export_kardex_pdf(kardex_entries: List[Dict[str, Any]], producto_nombre: str, codigo_barras: str, sucursal_nombre: str) -> bytes:
        """
        Genera el PDF del Kardex Valorado de un producto específico.
        """
        html_styles = ExportService.get_pdf_styles()
        
        # Generar las filas del reporte
        rows_html = ""
        total_entradas = Decimal("0.00")
        total_salidas = Decimal("0.00")

        for entry in kardex_entries:
            # Determinar tipo de badge
            tipo = entry["tipo_movimiento"]
            badge_class = "badge-ajuste"
            if "INGRESO" in tipo:
                badge_class = "badge-ingreso"
                total_entradas += Decimal(str(entry["cantidad_entrada"]))
            elif "SALIDA" in tipo:
                badge_class = "badge-salida"
                total_salidas += Decimal(str(entry["cantidad_salida"]))
            elif "TRASLADO" in tipo:
                badge_class = "badge-traslado"

            fecha_formateada = datetime.fromisoformat(str(entry["creado_en"])).strftime("%d/%m/%Y %H:%M") if entry.get("creado_en") else "N/A"

            # Formatear números
            cant_ent = f"{entry['cantidad_entrada']:.2f}" if entry["cantidad_entrada"] > 0 else "-"
            cost_ent = f"C$ {entry['costo_entrada']:.2f}" if entry["cantidad_entrada"] > 0 else "-"
            cant_sal = f"{entry['cantidad_salida']:.2f}" if entry["cantidad_salida"] > 0 else "-"
            cost_sal = f"C$ {entry['costo_salida']:.2f}" if entry["cantidad_salida"] > 0 else "-"
            
            rows_html += f"""
                <tr>
                    <td>{fecha_formateada}</td>
                    <td><span class="badge {badge_class}">{tipo}</span></td>
                    <td class="text-right">{cant_ent}</td>
                    <td class="text-right">{cost_ent}</td>
                    <td class="text-right">{cant_sal}</td>
                    <td class="text-right">{cost_sal}</td>
                    <td class="text-right font-bold">{entry['cantidad_saldo']:.2f}</td>
                    <td class="text-right">C$ {entry['costo_unitario_saldo']:.4f}</td>
                    <td class="text-right font-bold">C$ {entry['costo_total_saldo']:.2f}</td>
                </tr>
            """

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                {html_styles}
            </style>
        </head>
        <body>
            <div class="header-container">
                <table class="header-table">
                    <tr>
                        <td>
                            <div class="company-name">ERP MULTISUCURSAL</div>
                            <div class="report-title">Kardex Valorado (CPP)</div>
                        </td>
                        <td class="meta-info">
                            <strong>Fecha Emisión:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}<br>
                            <strong>Sucursal:</strong> {sucursal_nombre}<br>
                            <strong>Moneda:</strong> NIO (Córdoba nicaragüense)
                        </td>
                    </tr>
                </table>
            </div>

            <div style="margin-bottom: 20px; background-color: #EBF8FF; border: 1px solid #BEE3F8; border-radius: 6px; padding: 10px;">
                <table style="width: 100%; border-collapse: collapse; font-size: 9.5pt;">
                    <tr>
                        <td style="width: 15%; font-weight: bold; color: #2B6CB0;">Producto:</td>
                        <td style="width: 50%; font-weight: bold;">{producto_nombre}</td>
                        <td style="width: 15%; font-weight: bold; color: #2B6CB0;">Código/SKU:</td>
                        <td style="width: 20%; font-weight: bold;">{codigo_barras}</td>
                    </tr>
                </table>
            </div>

            <table class="table-report">
                <thead>
                    <tr>
                        <th>Fecha/Hora</th>
                        <th>Operación</th>
                        <th class="text-right">Entr. Cant</th>
                        <th class="text-right">Entr. Costo</th>
                        <th class="text-right">Sal. Cant</th>
                        <th class="text-right">Sal. Costo</th>
                        <th class="text-right">Saldo Cant</th>
                        <th class="text-right">Saldo Unit</th>
                        <th class="text-right">Saldo Total</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>

            <div class="summary-section">
                <table class="summary-table">
                    <tr>
                        <td class="font-bold">Total Entradas:</td>
                        <td class="text-right">{total_entradas:.2f}</td>
                    </tr>
                    <tr>
                        <td class="font-bold">Total Salidas:</td>
                        <td class="text-right">{total_salidas:.2f}</td>
                    </tr>
                    <tr style="border-top: 1px solid #CBD5E0;">
                        <td class="font-bold" style="color: #2B6CB0;">Stock Actual:</td>
                        <td class="text-right font-bold" style="color: #2B6CB0;">
                            {kardex_entries[-1]['cantidad_saldo']:.2f; if len(kardex_entries) > 0 else '0.00'}
                        </td>
                    </tr>
                    <tr>
                        <td class="font-bold" style="color: #2B6CB0;">Costo Promedio:</td>
                        <td class="text-right font-bold" style="color: #2B6CB0;">
                            C$ {kardex_entries[-1]['costo_unitario_saldo']:.4f; if len(kardex_entries) > 0 else '0.0000'}
                        </td>
                    </tr>
                    <tr style="border-top: 2px double #CBD5E0;">
                        <td class="font-bold" style="color: #1A365D; font-size: 10pt;">Valoración Total:</td>
                        <td class="text-right font-bold" style="color: #1A365D; font-size: 10pt;">
                            C$ {kardex_entries[-1]['costo_total_saldo']:.2f; if len(kardex_entries) > 0 else '0.00'}
                        </td>
                    </tr>
                </table>
            </div>
        </body>
        </html>
        """
        
        # En Python, format inside f-strings con {} no acepta punto y coma ni comentarios, corrijamos la sintaxis:
        # Reemplazamos las expresiones ternarias por valores ya computados
        stock_act = f"{kardex_entries[-1]['cantidad_saldo']:.2f}" if len(kardex_entries) > 0 else "0.00"
        costo_prom = f"{kardex_entries[-1]['costo_unitario_saldo']:.4f}" if len(kardex_entries) > 0 else "0.0000"
        val_tot = f"{kardex_entries[-1]['costo_total_saldo']:.2f}" if len(kardex_entries) > 0 else "0.00"

        html_content = html_content.replace("{kardex_entries[-1]['cantidad_saldo']:.2f; if len(kardex_entries) > 0 else '0.00'}", stock_act)
        html_content = html_content.replace("C$ {kardex_entries[-1]['costo_unitario_saldo']:.4f; if len(kardex_entries) > 0 else '0.0000'}", f"C$ {costo_prom}")
        html_content = html_content.replace("C$ {kardex_entries[-1]['costo_total_saldo']:.2f; if len(kardex_entries) > 0 else '0.00'}", f"C$ {val_tot}")

        # Generar PDF en memoria
        pdf_file = BytesIO()
        HTML(string=html_content).write_pdf(target=pdf_file)
        return pdf_file.getvalue()

    @staticmethod
    def export_accounting_pdf(balances: List[Dict[str, Any]]) -> bytes:
        """
        Genera el PDF del Balance de Comprobación (Plan de Cuentas Contable).
        """
        html_styles = ExportService.get_pdf_styles()
        
        rows_html = ""
        sum_debe = Decimal("0.00")
        sum_haber = Decimal("0.00")

        for entry in balances:
            # Solo acumulamos la suma total para las cuentas de nivel superior o nivel de detalle raíz?
            # En un balance de comprobación normal, la suma de débitos y créditos de todas las cuentas de detalle
            # (es_detalle == True) debe ser exactamente igual.
            if entry.get("es_detalle"):
                sum_debe += Decimal(str(entry["debe"]))
                sum_haber += Decimal(str(entry["haber"]))

            # Sangría visual según el nivel del código para simular jerarquía
            nivel = len(entry["codigo"].split("."))
            padding_left = (nivel - 1) * 12
            nombre_styled = f'<span style="padding-left: {padding_left}px; font-weight: {"bold" if not entry["es_detalle"] else "normal"}">{entry["nombre"]}</span>'
            codigo_styled = f'<span style="font-weight: {"bold" if not entry["es_detalle"] else "normal"}">{entry["codigo"]}</span>'

            debe_str = f"C$ {entry['debe']:.2f}" if entry['debe'] > 0 else "-"
            haber_str = f"C$ {entry['haber']:.2f}" if entry['haber'] > 0 else "-"
            saldo_str = f"C$ {entry['saldo']:.2f}" if entry['saldo'] != 0 else "C$ 0.00"

            rows_html += f"""
                <tr>
                    <td>{codigo_styled}</td>
                    <td>{nombre_styled}</td>
                    <td class="text-center">{entry['tipo']}</td>
                    <td class="text-center">{entry['naturaleza']}</td>
                    <td class="text-right">{debe_str}</td>
                    <td class="text-right">{haber_str}</td>
                    <td class="text-right font-bold">{saldo_str}</td>
                </tr>
            """

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                {html_styles}
            </style>
        </head>
        <body>
            <div class="header-container">
                <table class="header-table">
                    <tr>
                        <td>
                            <div class="company-name">ERP MULTISUCURSAL</div>
                            <div class="report-title">Balance de Comprobación Contable</div>
                        </td>
                        <td class="meta-info">
                            <strong>Fecha Emisión:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}<br>
                            <strong>Sucursal:</strong> Consolidated Global<br>
                            <strong>Moneda:</strong> NIO (Córdoba nicaragüense)
                        </td>
                    </tr>
                </table>
            </div>

            <table class="table-report">
                <thead>
                    <tr>
                        <th style="width: 12%;">Código</th>
                        <th style="width: 38%;">Cuenta Contable</th>
                        <th style="width: 10%; text-align: center;">Tipo</th>
                        <th style="width: 10%; text-align: center;">Nat.</th>
                        <th class="text-right" style="width: 10%;">Debe (Cargo)</th>
                        <th class="text-right" style="width: 10%;">Haber (Abono)</th>
                        <th class="text-right" style="width: 10%;">Saldo</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                    <tr style="background-color: #E2E8F0 !important; font-weight: bold; border-top: 2px solid #718096; border-bottom: 2px solid #718096;">
                        <td colspan="4">SUMAS DE CUENTAS DE DETALLE (PARTIDA DOBLE)</td>
                        <td class="text-right">C$ {sum_debe:.2f}</td>
                        <td class="text-right">C$ {sum_haber:.2f}</td>
                        <td class="text-right">Diferencia: C$ {(sum_debe - sum_haber).copy_abs():.2f}</td>
                    </tr>
                </tbody>
            </table>
        </body>
        </html>
        """

        pdf_file = BytesIO()
        HTML(string=html_content).write_pdf(target=pdf_file)
        return pdf_file.getvalue()

    @staticmethod
    def export_ventas_pdf(ventas_data: list, periodo_str: str = "") -> bytes:
        """Genera el PDF del Reporte de Ventas (Productos y Servicios)."""
        from decimal import Decimal
        html_styles = ExportService.get_pdf_styles()

        rows_html = ""
        total_general = Decimal("0.00")

        for row in ventas_data:
            subtotal = Decimal(str(row["subtotal"]))
            total_general += subtotal
            tipo = row["tipo_item"]
            tipo_badge = (
                '<span style="background:#EBF8FF;color:#2B6CB0;padding:2px 6px;border-radius:4px;font-size:7.5pt;font-weight:bold;">PRODUCTO</span>'
                if tipo == "PRODUCTO" else
                '<span style="background:#F0FFF4;color:#276749;padding:2px 6px;border-radius:4px;font-size:7.5pt;font-weight:bold;">SERVICIO</span>'
            )
            fecha_str = row["fecha"].strftime("%d/%m/%Y %H:%M") if hasattr(row["fecha"], "strftime") else str(row["fecha"])[:16]

            rows_html += f"""
                <tr>
                    <td>{fecha_str}</td>
                    <td style="font-weight:bold;">{row['numero_factura']}</td>
                    <td>{row.get('cliente','')}</td>
                    <td style="font-family:monospace;font-size:8pt;">{row['codigo_barras']}</td>
                    <td>{row['nombre_producto']}</td>
                    <td class="text-center">{tipo_badge}</td>
                    <td class="text-right">{float(row['cantidad']):.2f}</td>
                    <td class="text-right">C$ {float(row['precio_venta']):.2f}</td>
                    <td class="text-right font-bold">C$ {float(subtotal):.2f}</td>
                </tr>
            """

        periodo_info = f"<strong>Período:</strong> {periodo_str}" if periodo_str else f"<strong>Generado:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>{html_styles}</style>
        </head>
        <body>
            <div class="header-container">
                <table class="header-table">
                    <tr>
                        <td>
                            <div class="company-name">ERP SISTEMA</div>
                            <div class="report-title">Reporte de Ventas</div>
                        </td>
                        <td class="meta-info">
                            {periodo_info}<br>
                            <strong>Moneda:</strong> NIO (Córdoba nicaragüense)
                        </td>
                    </tr>
                </table>
            </div>

            <table class="table-report">
                <thead>
                    <tr>
                        <th>Fecha</th>
                        <th>No. Factura</th>
                        <th>Cliente</th>
                        <th>Código</th>
                        <th>Producto / Servicio</th>
                        <th class="text-center">Tipo</th>
                        <th class="text-right">Cantidad</th>
                        <th class="text-right">P. Venta</th>
                        <th class="text-right">Subtotal</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                    <tr style="background-color:#EBF8FF;font-weight:bold;border-top:2px solid #2B6CB0;">
                        <td colspan="8" style="text-align:right;padding-right:12px;">TOTAL GENERAL:</td>
                        <td class="text-right">C$ {float(total_general):.2f}</td>
                    </tr>
                </tbody>
            </table>
        </body>
        </html>
        """
        pdf_file = BytesIO()
        HTML(string=html_content).write_pdf(target=pdf_file)
        return pdf_file.getvalue()

    @staticmethod
    def export_ventas_excel(ventas_data: list) -> bytes:
        """Genera el Excel (.xlsx) del Reporte de Ventas."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from decimal import Decimal

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"

        # Estilos
        header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        total_fill = PatternFill(start_color="EBF8FF", end_color="EBF8FF", fill_type="solid")
        total_font = Font(bold=True, color="1A365D", size=10)
        thin_border = Border(
            left=Side(style="thin", color="CBD5E0"),
            right=Side(style="thin", color="CBD5E0"),
            top=Side(style="thin", color="CBD5E0"),
            bottom=Side(style="thin", color="CBD5E0"),
        )

        # Título
        ws.merge_cells("A1:I1")
        ws["A1"] = "REPORTE DE VENTAS"
        ws["A1"].font = Font(bold=True, size=14, color="1A365D")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:I2")
        ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font = Font(italic=True, size=9, color="718096")
        ws["A2"].alignment = Alignment(horizontal="center")

        # Headers
        headers = ["Fecha", "No. Factura", "Cliente", "Código", "Producto/Servicio", "Tipo", "Cantidad", "Precio Venta", "Subtotal"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Datos
        total_general = Decimal("0.00")
        for row_idx, row in enumerate(ventas_data, 5):
            subtotal = Decimal(str(row["subtotal"]))
            total_general += subtotal
            fecha_str = row["fecha"].strftime("%d/%m/%Y %H:%M") if hasattr(row["fecha"], "strftime") else str(row["fecha"])[:16]

            data_row = [
                fecha_str,
                row["numero_factura"],
                row.get("cliente", ""),
                row["codigo_barras"],
                row["nombre_producto"],
                row["tipo_item"],
                float(row["cantidad"]),
                float(row["precio_venta"]),
                float(subtotal),
            ]
            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                cell.border = thin_border
                if col_num in [7, 8, 9]:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")

        # Totales
        total_row = len(ventas_data) + 5
        ws.merge_cells(f"A{total_row}:H{total_row}")
        ws[f"A{total_row}"] = "TOTAL GENERAL:"
        ws[f"A{total_row}"].font = total_font
        ws[f"A{total_row}"].fill = total_fill
        ws[f"A{total_row}"].alignment = Alignment(horizontal="right")

        total_cell = ws.cell(row=total_row, column=9, value=float(total_general))
        total_cell.font = total_font
        total_cell.fill = total_fill
        total_cell.number_format = '#,##0.00'
        total_cell.alignment = Alignment(horizontal="right")

        # Anchos de columna
        col_widths = [18, 14, 20, 14, 30, 12, 10, 12, 14]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
